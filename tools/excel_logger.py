from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

FILE_NAME = "leads.xlsx"


def save_to_excel(lead, status, action):
    """
    Saves outreach activity into Excel.
    Creates file automatically if missing.
    """

    # ================= OPEN / CREATE FILE =================
    if os.path.exists(FILE_NAME):
        wb = load_workbook(FILE_NAME)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

        # 🔥 HEADER
        ws.append([
            "Name",
            "Email",
            "LinkedIn",
            "Website",
            "Status",
            "Action",
            "Timestamp"
        ])

    # ================= ADD ROW =================
    ws.append([
        lead.get("name", ""),
        lead.get("outreach_email", ""),
        lead.get("linkedin_url", ""),
        lead.get("website", ""),
        status,
        action,
        datetime.now().strftime("%Y-%m-%d %H:%M")
    ])

    # ================= AUTO COLUMN WIDTH =================
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = min(length + 5, 50)

        column_letter = column_cells[0].column_letter
        ws.column_dimensions[column_letter].width = adjusted_width

    # ================= SAVE FILE =================
    wb.save(FILE_NAME)

    print(f"📊 Excel updated → {FILE_NAME}")